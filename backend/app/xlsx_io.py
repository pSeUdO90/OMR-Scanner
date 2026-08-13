from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook

HEADER_ALIASES = {
    "roll_no": {"roll no", "roll", "rollno", "roll_no", "roll number"},
    "name": {"student name", "name", "candidate name"},
    "gender": {"gender", "sex"},
    "class_name": {"class", "class_name", "class name", "std"},
    "section": {"section", "sec"},
    "session": {"session", "academic session", "year"},
}


def _norm(value: object) -> str:
    return str(value or "").strip()


def parse_students_xlsx(content: bytes) -> list[dict[str, str]]:
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_norm(h).lower() for h in rows[0]]
    index: dict[str, int] = {}
    for i, header in enumerate(headers):
        for field, aliases in HEADER_ALIASES.items():
            if header in aliases:
                index[field] = i
    if "roll_no" not in index or "name" not in index:
        raise ValueError("XLSX must include Roll No and Student Name columns.")
    students = []
    for row in rows[1:]:
        roll = _norm(row[index["roll_no"]] if index["roll_no"] < len(row) else "")
        name = _norm(row[index["name"]] if index["name"] < len(row) else "")
        if not roll or not name:
            continue
        students.append(
            {
                "roll_no": roll,
                "name": name,
                "gender": _norm(row[index["gender"]]) if "gender" in index else "",
                "class_name": _norm(row[index["class_name"]]) if "class_name" in index else "",
                "section": _norm(row[index["section"]]) if "section" in index else "",
                "session": _norm(row[index["session"]]) if "session" in index else "",
            }
        )
    return students


def students_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(["Roll No", "Student Name", "Gender", "Class", "Section", "Session"])
    ws.append(["2400100001", "Aarav Mishra", "M", "12", "A", "2025-26"])
    ws.append(["2400100002", "Diya Patnaik", "F", "12", "A", "2025-26"])
    ws.append(["2400100003", "Rohan Das", "M", "12", "B", "2025-26"])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
