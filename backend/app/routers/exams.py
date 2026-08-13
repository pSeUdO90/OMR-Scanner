from __future__ import annotations

import csv
import json
import shutil
from io import BytesIO, StringIO
from pathlib import Path
from uuid import uuid4

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from PIL import Image
from sqlalchemy.orm import Session, joinedload

from ..database import UPLOAD_DIR, get_db
from ..models import Exam, ExamSheet, ExamSubjectMap, OmrLayout, Student, Subject
from ..omr.analyze import analyze_layout_config
from ..omr.generator import generate_sheet, prefill_on_layout_sample
from ..omr.processor import evaluate_image, load_image, parse_layout, save_image
from ..omr.sample_file import sample_to_image_bytes
from ..schemas import AnswerKeyIn, AssignSheetIn, ExamIn, ExamOut, GraceIn, SheetOut, SubjectMapOut
from ..scoring import assigned_students, bind_sheet_student, build_analytics, parse_question_numbers, rescore_stored_sheets, score_sheet

router = APIRouter(prefix="/api/exams", tags=["exams"])


def allocate_test_id(db: Session) -> str:
    values = [row[0] for row in db.query(Exam.test_id).all() if row[0]]
    highest = 0
    for value in values:
        digits = "".join(ch for ch in str(value) if ch.isdigit())
        if digits:
            highest = max(highest, int(digits))
    return f"{highest + 1:04d}"


def _exam_out(exam: Exam, *, with_analysis: bool = False) -> ExamOut:
    maps = []
    for mapping in exam.subject_maps:
        maps.append(
            SubjectMapOut(
                id=mapping.id,
                subject_id=mapping.subject_id,
                start_q=mapping.start_q,
                end_q=mapping.end_q,
                subject_name=mapping.subject.name if mapping.subject else "",
            )
        )
    sheets = exam.sheets or []
    return ExamOut(
        id=exam.id,
        name=exam.name,
        exam_date=exam.exam_date,
        exam_type=exam.exam_type,
        duration_minutes=exam.duration_minutes,
        correct_marks=exam.correct_marks,
        wrong_marks=exam.wrong_marks,
        unattempted_marks=exam.unattempted_marks,
        layout_id=exam.layout_id,
        layout_name=exam.layout.name if exam.layout else "",
        total_questions=exam.layout.total_questions if exam.layout else 0,
        status=exam.status,
        created_at=exam.created_at,
        subject_maps=maps,
        answer_key=json.loads(exam.answer_key_json or "{}"),
        sheet_count=len(sheets),
        evaluated_count=sum(1 for s in sheets if s.status in ("evaluated", "unmatched")),
        has_sample=bool(getattr(exam, "sample_path", "")),
        test_id=getattr(exam, "test_id", "") or "",
        test_no=getattr(exam, "test_no", "") or "",
        class_name=getattr(exam, "class_name", "") or "",
        section=getattr(exam, "section", "") or "",
        batch=getattr(exam, "batch", "") or "",
        grace_questions=json.loads(getattr(exam, "grace_questions_json", None) or "[]"),
        field_map=json.loads(getattr(exam, "field_map_json", None) or "{}")
        or json.loads(getattr(exam.layout, "field_map_json", None) or "{}"),
        analysis=_exam_analysis(exam) if with_analysis else [],
    )


def _exam_analysis(exam: Exam) -> list[dict]:
    if not exam.layout:
        return []
    layout = parse_layout(exam.layout.config_json)
    img = None
    if getattr(exam, "sample_path", "") and Path(exam.sample_path).exists():
        try:
            img = load_image(exam.sample_path)
        except Exception:
            img = None
    return analyze_layout_config(layout, img)


def _load_exam(db: Session, exam_id: int) -> Exam:
    exam = (
        db.query(Exam)
        .options(
            joinedload(Exam.layout),
            joinedload(Exam.subject_maps).joinedload(ExamSubjectMap.subject),
            joinedload(Exam.sheets).joinedload(ExamSheet.student),
            joinedload(Exam.sheets).joinedload(ExamSheet.question_results),
        )
        .filter(Exam.id == exam_id)
        .one_or_none()
    )
    if not exam:
        raise HTTPException(404, "Exam not found")
    return exam


@router.get("", response_model=list[ExamOut])
def list_exams(db: Session = Depends(get_db)):
    exams = db.query(Exam).options(joinedload(Exam.layout), joinedload(Exam.sheets), joinedload(Exam.subject_maps).joinedload(ExamSubjectMap.subject)).order_by(Exam.id.desc()).all()
    return [_exam_out(exam) for exam in exams]


@router.get("/next-test-id")
def get_next_test_id(db: Session = Depends(get_db)):
    return {"test_id": allocate_test_id(db)}


@router.post("", response_model=ExamOut)
def create_exam(payload: ExamIn, db: Session = Depends(get_db)):
    layout = db.get(OmrLayout, payload.layout_id)
    if not layout:
        raise HTTPException(400, "Layout not found")
    exam = Exam(
        name=payload.name,
        exam_date=payload.exam_date,
        exam_type=payload.exam_type,
        duration_minutes=payload.duration_minutes,
        correct_marks=payload.correct_marks,
        wrong_marks=payload.wrong_marks,
        unattempted_marks=payload.unattempted_marks,
        layout_id=payload.layout_id,
        answer_key_json=json.dumps(payload.answer_key),
        test_id=allocate_test_id(db),
        test_no=payload.test_no,
        class_name=payload.class_name,
        section=payload.section,
        batch=payload.batch,
        status="draft",
    )
    db.add(exam)
    db.flush()
    _replace_subject_maps(db, exam, payload.subject_maps, layout)
    db.commit()
    return _exam_out(_load_exam(db, exam.id))


def _replace_subject_maps(db: Session, exam: Exam, payload_maps: list, layout: OmrLayout) -> None:
    maps = [
        {"subject_id": m.subject_id, "start_q": m.start_q, "end_q": m.end_q}
        for m in payload_maps
    ]
    if not maps:
        preview = json.loads(layout.config_json).get("default_maps", [])
        by_name = {s.name: s for s in db.query(Subject).all()}
        for item in preview:
            subject = by_name.get(item["subject"])
            if subject:
                maps.append(
                    {
                        "subject_id": subject.id,
                        "start_q": item["start_q"],
                        "end_q": item["end_q"],
                    }
                )
    db.query(ExamSubjectMap).filter(ExamSubjectMap.exam_id == exam.id).delete()
    db.flush()
    for mapping in maps:
        db.add(
            ExamSubjectMap(
                exam_id=exam.id,
                subject_id=mapping["subject_id"],
                start_q=mapping["start_q"],
                end_q=mapping["end_q"],
            )
        )


@router.get("/{exam_id}", response_model=ExamOut)
def get_exam(exam_id: int, db: Session = Depends(get_db)):
    return _exam_out(_load_exam(db, exam_id), with_analysis=True)


@router.put("/{exam_id}", response_model=ExamOut)
def update_exam(exam_id: int, payload: ExamIn, db: Session = Depends(get_db)):
    exam = _load_exam(db, exam_id)
    if exam.status in ("evaluated", "published"):
        raise HTTPException(409, "Exam already evaluated. Only the answer key can be changed.")
    layout = db.get(OmrLayout, payload.layout_id)
    if not layout:
        raise HTTPException(400, "Layout not found")
    exam.name = payload.name
    exam.exam_date = payload.exam_date
    exam.exam_type = payload.exam_type
    exam.duration_minutes = payload.duration_minutes
    exam.correct_marks = payload.correct_marks
    exam.wrong_marks = payload.wrong_marks
    exam.unattempted_marks = payload.unattempted_marks
    exam.layout_id = payload.layout_id
    exam.test_no = payload.test_no
    exam.class_name = payload.class_name
    exam.section = payload.section
    exam.batch = payload.batch
    _replace_subject_maps(db, exam, payload.subject_maps, layout)
    db.commit()
    return _exam_out(_load_exam(db, exam.id))


@router.put("/{exam_id}/grace", response_model=ExamOut)
def set_grace_marks(exam_id: int, payload: GraceIn, db: Session = Depends(get_db)):
    exam = _load_exam(db, exam_id)
    exam.grace_questions_json = json.dumps(parse_question_numbers(payload.questions))
    rescore_stored_sheets(db, exam)
    db.commit()
    return _exam_out(_load_exam(db, exam.id))


@router.delete("/{exam_id}")
def delete_exam(exam_id: int, db: Session = Depends(get_db)):
    exam = db.get(Exam, exam_id)
    if not exam:
        raise HTTPException(404, "Exam not found")
    dest = UPLOAD_DIR / f"exam-{exam.id}"
    db.delete(exam)
    db.commit()
    if dest.exists():
        shutil.rmtree(dest, ignore_errors=True)
    return {"ok": True}


@router.post("/{exam_id}/sample")
async def upload_omr_sample(exam_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    exam = _load_exam(db, exam_id)
    dest_dir = UPLOAD_DIR / f"exam-{exam.id}"
    dest_dir.mkdir(parents=True, exist_ok=True)
    raw = await file.read()
    try:
        image_bytes, suffix = sample_to_image_bytes(file.filename or "sample.jpg", raw)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    stored = dest_dir / f"omr-sample{suffix}"
    stored.write_bytes(image_bytes)
    exam.sample_path = str(stored)
    db.commit()
    layout = parse_layout(exam.layout.config_json)
    analysis = analyze_layout_config(layout, load_image(stored))
    return {
        "ok": True,
        "filename": file.filename,
        "analysis": analysis,
        "field_map": json.loads(
            getattr(exam, "field_map_json", None)
            or getattr(exam.layout, "field_map_json", None)
            or "{}"
        ),
        "targets": [
            {"value": "", "label": "Ignore"},
            {"value": "exam_date", "label": "Exam Date"},
            {"value": "test_id", "label": "Test ID"},
            {"value": "test_no", "label": "Test No"},
        ],
    }


@router.post("/{exam_id}/field-map")
def save_exam_field_map(exam_id: int, payload: dict, db: Session = Depends(get_db)):
    exam = _load_exam(db, exam_id)
    mapping = payload.get("field_map") or payload
    exam.field_map_json = json.dumps(mapping)
    db.commit()
    return {"ok": True, "field_map": mapping}


@router.get("/{exam_id}/sample")
def get_omr_sample(exam_id: int, db: Session = Depends(get_db)):
    exam = _load_exam(db, exam_id)
    if not exam.sample_path or not Path(exam.sample_path).exists():
        raise HTTPException(404, "No OMR sample uploaded")
    return FileResponse(exam.sample_path)


@router.post("/{exam_id}/answer-key/upload", response_model=ExamOut)
async def upload_answer_key(exam_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    exam = _load_exam(db, exam_id)
    raw = await file.read()
    name = (file.filename or "").lower()
    key: dict[str, str] = {}
    if name.endswith((".png", ".jpg", ".jpeg", ".tif", ".tiff", ".webp", ".bmp")):
        dest = UPLOAD_DIR / f"exam-{exam.id}" / f"key-{uuid4().hex}{Path(name).suffix or '.png'}"
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(raw)
        layout = parse_layout(exam.layout.config_json)
        result = evaluate_image(load_image(dest), layout)
        key = {q: ans for q, ans in result["answers"].items() if ans and ans != "MULTI"}
    else:
        text = raw.decode("utf-8", errors="ignore")
        letters = [ch.upper() for ch in text if ch.upper() in "ABCD"]
        if "," in text or "\t" in text:
            key = {}
            for line in text.splitlines():
                parts = [p.strip() for p in line.replace("\t", ",").split(",") if p.strip()]
                if len(parts) >= 2 and parts[0].isdigit() and parts[1].upper()[:1] in "ABCD":
                    key[str(int(parts[0]))] = parts[1].upper()[:1]
        if not key:
            key = {str(i + 1): letter for i, letter in enumerate(letters)}
    if not key:
        raise HTTPException(400, "Could not read an answer key from that file")
    exam.answer_key_json = json.dumps(key)
    rescore_stored_sheets(db, exam)
    db.commit()
    return _exam_out(_load_exam(db, exam.id))


@router.put("/{exam_id}/answer-key", response_model=ExamOut)
def set_answer_key(exam_id: int, payload: AnswerKeyIn, db: Session = Depends(get_db)):
    exam = _load_exam(db, exam_id)
    key = payload.answer_key or {}
    if payload.key_string:
        letters = [ch.upper() for ch in payload.key_string if ch.upper() in "ABCD"]
        key = {str(i + 1): letter for i, letter in enumerate(letters)}
    exam.answer_key_json = json.dumps(key)
    rescore_stored_sheets(db, exam)
    db.commit()
    return _exam_out(_load_exam(db, exam_id))


@router.post("/{exam_id}/sheets", response_model=list[SheetOut])
async def upload_sheets(exam_id: int, files: list[UploadFile] = File(...), db: Session = Depends(get_db)):
    exam = _load_exam(db, exam_id)
    saved = []
    dest_dir = UPLOAD_DIR / f"exam-{exam.id}"
    dest_dir.mkdir(parents=True, exist_ok=True)
    for upload in files:
        suffix = Path(upload.filename or "sheet.png").suffix or ".png"
        stored = dest_dir / f"{uuid4().hex}{suffix}"
        stored.write_bytes(await upload.read())
        sheet = ExamSheet(
            exam_id=exam.id,
            filename=upload.filename or stored.name,
            stored_path=str(stored),
            status="uploaded",
        )
        db.add(sheet)
        db.commit()
        db.refresh(sheet)
        try:
            layout = parse_layout(exam.layout.config_json)
            result = evaluate_image(load_image(stored), layout)
            bind_sheet_student(db, exam, sheet, result["roll"])
            db.commit()
            db.refresh(sheet)
        except Exception:
            pass
        saved.append(_sheet_out(sheet))
    return saved


def _sheet_out(sheet: ExamSheet) -> SheetOut:
    return SheetOut(
        id=sheet.id,
        exam_id=sheet.exam_id,
        student_id=sheet.student_id,
        student_name=sheet.student.name if sheet.student else "",
        filename=sheet.filename,
        status=sheet.status,
        detected_roll=sheet.detected_roll,
        error_message=sheet.error_message,
        raw_score=sheet.raw_score,
        max_score=sheet.max_score,
        right_count=sheet.right_count,
        wrong_count=sheet.wrong_count,
        left_count=sheet.left_count,
        invalid_count=sheet.invalid_count,
        has_overlay=bool(sheet.overlay_path and Path(sheet.overlay_path).exists()),
        assigned_manually=bool(getattr(sheet, "assigned_manually", False)),
    )


@router.get("/{exam_id}/sheets", response_model=list[SheetOut])
def list_sheets(exam_id: int, db: Session = Depends(get_db)):
    exam = _load_exam(db, exam_id)
    return [_sheet_out(sheet) for sheet in exam.sheets]


@router.post("/{exam_id}/evaluate")
def evaluate_exam(exam_id: int, db: Session = Depends(get_db)):
    exam = _load_exam(db, exam_id)
    key = json.loads(exam.answer_key_json or "{}")
    if not key:
        raise HTTPException(400, "Set an answer key before evaluating")
    layout = parse_layout(exam.layout.config_json)
    evaluated = 0
    errors = []
    for sheet in exam.sheets:
        try:
            image = load_image(sheet.stored_path)
            result = evaluate_image(image, layout)
            overlay_path = str(Path(sheet.stored_path).with_name(f"overlay-{sheet.id}.png"))
            save_image(overlay_path, result["overlay"])
            sheet.overlay_path = overlay_path
            score_sheet(db, exam, sheet, result["answers"], result["roll"])
            evaluated += 1
        except Exception as exc:  # noqa: BLE001
            sheet.status = "error"
            sheet.error_message = str(exc)
            errors.append({"sheet_id": sheet.id, "error": str(exc)})
    if exam.status != "published":
        exam.status = "evaluated"
    db.commit()
    return {"evaluated": evaluated, "errors": errors}


@router.post("/{exam_id}/publish")
def publish_exam(exam_id: int, db: Session = Depends(get_db)):
    exam = _load_exam(db, exam_id)
    if exam.status not in ("evaluated", "published"):
        raise HTTPException(400, "Evaluate sheets before publishing")
    exam.status = "published"
    db.commit()
    return {"status": "published"}


@router.get("/{exam_id}/results")
def exam_results(exam_id: int, db: Session = Depends(get_db)):
    exam = _load_exam(db, exam_id)
    return build_analytics(exam)


@router.get("/{exam_id}/results.csv")
def exam_results_csv(exam_id: int, db: Session = Depends(get_db)):
    exam = _load_exam(db, exam_id)
    analytics = build_analytics(exam)
    buf = StringIO()
    writer = csv.writer(buf)
    subject_names = [s["subject_name"] for s in analytics["subjects"]]
    writer.writerow(
        ["Rank", "Roll No", "Name", "Class", "Section", "Right", "Wrong", "Left", "Invalid", "Score", "Max", "Percentage"]
        + [f"{name} R" for name in subject_names]
        + [f"{name} W" for name in subject_names]
        + [f"{name} L" for name in subject_names]
    )
    for row in analytics["results"]:
        by_subject = {s["subject_name"]: s for s in row["subjects"]}
        writer.writerow(
            [
                row["rank"],
                row["roll_no"],
                row["name"],
                row["class_name"],
                row["section"],
                row["right"],
                row["wrong"],
                row["left"],
                row["invalid"],
                row["score"],
                row["max_score"],
                row["percentage"],
            ]
            + [by_subject.get(name, {}).get("right", "") for name in subject_names]
            + [by_subject.get(name, {}).get("wrong", "") for name in subject_names]
            + [by_subject.get(name, {}).get("left", "") for name in subject_names]
        )
    data = buf.getvalue().encode("utf-8")
    filename = f"{exam.name.replace(' ', '_')}_rwl.csv"
    return StreamingResponse(
        iter([data]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{exam_id}/results.xlsx")
def exam_results_xlsx(exam_id: int, db: Session = Depends(get_db)):
    exam = _load_exam(db, exam_id)
    analytics = build_analytics(exam)
    wb = Workbook()
    ranks = wb.active
    ranks.title = "Rank list"
    subject_names = [s["subject_name"] for s in analytics["subjects"]]
    headers = ["Rank", "Roll No", "Name", "Class", "Section", "Right", "Wrong", "Left", "Invalid", "Score", "Max", "Percentage"]
    headers += [f"{name} R" for name in subject_names]
    headers += [f"{name} W" for name in subject_names]
    headers += [f"{name} L" for name in subject_names]
    ranks.append(headers)
    for cell in ranks[1]:
        cell.font = Font(bold=True)
    for row in analytics["results"]:
        by_subject = {s["subject_name"]: s for s in row["subjects"]}
        ranks.append(
            [
                row["rank"],
                row["roll_no"],
                row["name"],
                row["class_name"],
                row["section"],
                row["right"],
                row["wrong"],
                row["left"],
                row["invalid"],
                row["score"],
                row["max_score"],
                row["percentage"],
            ]
            + [by_subject.get(name, {}).get("right", "") for name in subject_names]
            + [by_subject.get(name, {}).get("wrong", "") for name in subject_names]
            + [by_subject.get(name, {}).get("left", "") for name in subject_names]
        )

    overall = wb.create_sheet("Overall RWL")
    overall.append(["Exam", analytics["exam_name"]])
    overall.append(["Appeared", analytics["appeared"]])
    overall.append(["Average", analytics["average_score"]])
    overall.append(["Highest", analytics["highest_score"]])
    overall.append(["Lowest", analytics["lowest_score"]])
    overall.append([])
    overall.append(["Subject", "Right", "Wrong", "Left", "Invalid", "Accuracy", "Score", "Max"])
    for cell in overall[7]:
        cell.font = Font(bold=True)
    rwl_rows = [analytics["overall_rwl"], *analytics["subjects"]]
    for item in rwl_rows:
        overall.append(
            [item["subject_name"], item["right"], item["wrong"], item["left"], item["invalid"], item["accuracy"], item["score"], item["max_score"]]
        )

    items = wb.create_sheet("Item analysis")
    items.append(["Question", "Key", "Right", "Wrong", "Left", "Invalid", "Difficulty"])
    for cell in items[1]:
        cell.font = Font(bold=True)
    for item in analytics["item_analysis"]:
        items.append([item["question_no"], item["correct"], item["right"], item["wrong"], item["left"], item["invalid"], item["difficulty"]])

    buf = BytesIO()
    wb.save(buf)
    filename = f"{exam.name.replace(' ', '_')}_rwl.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _get_sheet(db: Session, exam_id: int, sheet_id: int) -> ExamSheet:
    sheet = db.get(ExamSheet, sheet_id)
    if not sheet or sheet.exam_id != exam_id:
        raise HTTPException(404, "Sheet not found")
    return sheet


@router.get("/{exam_id}/sheets/{sheet_id}/overlay")
def sheet_overlay(exam_id: int, sheet_id: int, db: Session = Depends(get_db)):
    sheet = _get_sheet(db, exam_id, sheet_id)
    if sheet.overlay_path and Path(sheet.overlay_path).exists():
        return FileResponse(sheet.overlay_path)
    if sheet.stored_path and Path(sheet.stored_path).exists():
        return FileResponse(sheet.stored_path)
    raise HTTPException(404, "Sheet image not found")


@router.get("/{exam_id}/sheets/{sheet_id}/image")
def sheet_image(exam_id: int, sheet_id: int, db: Session = Depends(get_db)):
    sheet = _get_sheet(db, exam_id, sheet_id)
    path = sheet.overlay_path if sheet.overlay_path and Path(sheet.overlay_path).exists() else sheet.stored_path
    if not path or not Path(path).exists():
        raise HTTPException(404, "Sheet image not found")
    return FileResponse(path)


@router.put("/{exam_id}/sheets/{sheet_id}/assign", response_model=SheetOut)
def assign_sheet(exam_id: int, sheet_id: int, payload: AssignSheetIn, db: Session = Depends(get_db)):
    exam = _load_exam(db, exam_id)
    sheet = _get_sheet(db, exam_id, sheet_id)
    student = db.get(Student, payload.student_id)
    if not student:
        raise HTTPException(404, "Student not found")
    sheet.student_id = student.id
    sheet.assigned_manually = True
    sheet.error_message = ""
    if sheet.answers_json and sheet.answers_json != "{}":
        sheet.status = "evaluated"
    elif sheet.status == "unmatched":
        sheet.status = "uploaded"
    db.commit()
    db.refresh(sheet)
    return _sheet_out(sheet)


@router.post("/{exam_id}/reset-omr")
def reset_omr(exam_id: int, db: Session = Depends(get_db)):
    exam = _load_exam(db, exam_id)
    removed = 0
    for sheet in list(exam.sheets):
        for path in (sheet.stored_path, sheet.overlay_path):
            if path:
                file_path = Path(path)
                if file_path.exists() and file_path.is_file():
                    file_path.unlink()
        db.delete(sheet)
        removed += 1
    if exam.status in ("evaluated", "published"):
        exam.status = "draft"
    db.commit()
    return {"ok": True, "removed": removed, "status": exam.status}


@router.post("/{exam_id}/sample-sheet")
def make_sample_sheet(
    exam_id: int,
    roll: str = Form(...),
    answers: str = Form(""),
    db: Session = Depends(get_db),
):
    exam = _load_exam(db, exam_id)
    layout = parse_layout(exam.layout.config_json)
    key = json.loads(exam.answer_key_json or "{}")
    parsed: dict[int, str] = {}
    if answers:
        letters = [ch.upper() for ch in answers if ch.upper() in "ABCD"]
        parsed = {i + 1: letter for i, letter in enumerate(letters)}
    elif key:
        parsed = {int(k): v for k, v in key.items()}
    image = generate_sheet(
        layout,
        roll,
        parsed,
        test_id=exam.test_id or "",
        test_no=exam.test_no or "",
    )
    dest = UPLOAD_DIR / f"exam-{exam.id}" / f"sample-{uuid4().hex}.png"
    save_image(dest, image)
    sheet = ExamSheet(exam_id=exam.id, filename=dest.name, stored_path=str(dest), status="uploaded")
    db.add(sheet)
    db.commit()
    db.refresh(sheet)
    return _sheet_out(sheet)


@router.get("/{exam_id}/prefilled-omr")
def prefilled_omr_pdf(exam_id: int, db: Session = Depends(get_db)):
    exam = _load_exam(db, exam_id)
    students = assigned_students(db, exam)
    if not students:
        raise HTTPException(400, "No students assigned to this exam. Set class, section, and batch from the student list.")
    layout = parse_layout(exam.layout.config_json)
    sample_path = getattr(exam.layout, "sample_path", "") or ""
    if not sample_path or not Path(sample_path).exists():
        raise HTTPException(400, "Upload an OMR layout sample PDF/JPG before generating pre-filled sheets.")
    base = load_image(sample_path)
    pages = []
    exam_date = str(exam.exam_date)
    for student in students:
        image = prefill_on_layout_sample(
            base,
            layout,
            roll=student.roll_no,
            student_name=student.name,
            test_id=exam.test_id or "",
            test_no=exam.test_no or "",
            exam_date=exam_date,
        )
        rgb = image[:, :, ::-1]
        pages.append(Image.fromarray(rgb))
    buf = BytesIO()
    pages[0].save(buf, format="PDF", save_all=True, append_images=pages[1:], resolution=120)
    filename = f"{exam.name.replace(' ', '_')}_prefilled_omr.pdf"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
